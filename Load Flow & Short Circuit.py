
# Connect to PowerFactory
import powerfactory as pf
import openpyxl as xl

app = pf.GetApplication()
app.ClearOutputWindow()

# Get active project
prj = app.GetActiveProject()
filename = prj.GetAttribute("loc_name")

# Get all relevant parameter
buses = app.GetCalcRelevantObjects('*.ElmTerm')

# Retrieve load-flow object
ldf = app.GetFromStudyCase("ComLdf")
# Force balanced load flow
ldf.iopt_net = 0

# Execute load flow
ldf.Execute()

# Create a new workbook for both load flow and short circuit data
wb = xl.Workbook()

# Create sheet for Power Flow data
sheet1 = wb.active
sheet1.title = 'Power Flow'

sheet1['A1'] = 'Bus'
sheet1['B1'] = 'Name'
sheet1['C1'] = 'V (p.u.)'
sheet1['D1'] = 'Angle (deg)'
sheet1['E1'] = 'P (MW)'
sheet1['F1'] = 'Q (MVar)'

i = 0

for bus in buses:
    i += 1
    Name = bus.GetAttribute("loc_name")
    V = bus.GetAttribute("m:u")
    Angle = bus.GetAttribute("m:phiu")
    P = bus.GetAttribute("m:Pflow")
    Q = bus.GetAttribute("m:Qflow")

    # Write to excel file
    sheet1.cell(row=1 + i, column=1).value = i
    sheet1.cell(row=1 + i, column=2).value = Name
    sheet1.cell(row=1 + i, column=3).value = V
    sheet1.cell(row=1 + i, column=4).value = Angle
    sheet1.cell(row=1 + i, column=5).value = P
    sheet1.cell(row=1 + i, column=6).value = Q

# Retrieve short circuit object
sc = app.GetFromStudyCase("ComShc")


  
# Execute short circuit
sc.Execute()

# Create sheet for Short Circuit data
sheet2 = wb.create_sheet(title='Short Circuit')

sheet2['A1'] = 'Bus'
sheet2['B1'] = 'Name'
sheet2['C1'] = 'SC Current (kA.)'
sheet2['D1'] = 'Peak Short Current (kA)'

g = 0

for bus in buses:
    g += 1
    Name = bus.GetAttribute("loc_name")
    I = bus.GetAttribute("m:Skss")
    Ps = bus.GetAttribute("m:Ikss")

    # Write to excel file
    sheet2.cell(row=1 + g, column=1).value = g
    sheet2.cell(row=1 + g, column=2).value = Name
    sheet2.cell(row=1 + g, column=3).value = I
    sheet2.cell(row=1 + g, column=4).value = Ps

# Save the excel file with both sheets
wb.save('C:\\Users\\User\\results\\%s.xlsx' % ('PowerFlow_ShortCircuit_Data'))
